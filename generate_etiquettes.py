import shutil
import openpyxl


if __name__ == "__main__":

    # Chemin du fichier source
    source_file = './TJX_ETIQUETTE.xlsx'

    wb_in = openpyxl.load_workbook('TJX_info.xlsx')
    ws_in = wb_in.active

    print("GENERATIONS DES ETIQUETTES")
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        print("|", end ="")
        ref_article, po, departement, nb_colis, colisage, pretickets, store_ready, lourd, fragile, taille, quantite = row[:11]

        # Copie du modèle
        destination_file = './Etiquettes/TJX_ETIQUETTE_'+str(ref_article)+'.xlsx'

        # Copie du fichier en utilisant la commande "cp"
        shutil.copy(source_file, destination_file)

        
        # Ouvrez le classeur existant
        wb = openpyxl.load_workbook(destination_file)


        # Sélectionnez la feuille de calcul
        ws = wb.active

        #po
        t = 'K2'

        for i in str(po):
            letter = t[0]
            num = t[1:]
            letter = chr(ord(letter) + 1)
            t = letter + num
            ws[t] = i

        # Dept
        ws['O4'] = departement

        # Nb Colis
        ws['R6'] = nb_colis

        # Colisage
        ws['O8'] = colisage

        # Pretickets
        if (pretickets == "Y" or pretickets == "y"):
            ws['P10'] = 'X'
        else :
            ws['S10'] = 'X'

        # Store Ready
        if (store_ready == "Y" or store_ready == "y"):
            ws['P12'] = 'X'
        else:
            ws['S12'] = 'X'

        # Heavy
        if (lourd == "Y" or lourd == "y"):
            ws['P14'] = 'X'
        else:
            ws['S14'] = 'X'

        # Fragile
        if (fragile == "Y" or fragile == "y"):
            ws['P16'] = 'X'
        else:
            ws['S16'] = 'X'

        # Colour

        # Vendor style
        ws['D18'] = str(ref_article)

        # Size
        t = 'C20'
        for i in str(taille):
            letter = t[0]
            num = t[1:]
            letter = chr(ord(letter) + 1)
            t = letter + num
            ws[t] = i

        # Quantity
        t = 'C22'
        for i in str(quantite):
            letter = t[0]
            num = t[1:]
            letter = chr(ord(letter) + 1)
            t = letter + num
            ws[t] = i
        
        # Enregistrez le fichier
        wb.save(destination_file)
    

    print()
    print("ETIQUETTES GENEREES DANS LE DOSSIER Etiquettes")